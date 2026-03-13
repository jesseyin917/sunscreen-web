import csv
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / 'data'
OUT = ROOT / 'data.js'

age_cols = [
    'Age_0_to_4','Age_5_to_9','Age_10_to_14','Age_15_to_19','Age_20_to_24','Age_25_to_29',
    'Age_30_to_34','Age_35_to_39','Age_40_to_44','Age_45_to_49','Age_50_to_54','Age_55_to_59',
    'Age_60_to_64','Age_65_to_69','Age_70_to_74','Age_75_to_79','Age_80_to_84','Age_85+'
]

# --- Cancer data (data.gov.au / AIHW) ---
cancer_rows = []
with open(DATA_DIR / 'acimcombinedcounts.csv', newline='', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        if row['Cancer_Type'] == 'Melanoma of the skin' and row['Sex'] == 'Persons':
            total = sum(int(row[c] or 0) for c in age_cols)
            if total > 0:
                cancer_rows.append({
                    'year': int(row['Year']),
                    'type': row['Type'],
                    'value': total,
                })

incidence = [r for r in cancer_rows if r['type'] == 'Incidence']
mortality = [r for r in cancer_rows if r['type'] == 'Mortality']

# keep overlapping non-zero period
inc_years = {r['year'] for r in incidence}
mort_years = {r['year'] for r in mortality}
years = sorted(inc_years & mort_years)
melanoma_trend = []
for y in years:
    melanoma_trend.append({
        'year': y,
        'incidence': next(r['value'] for r in incidence if r['year'] == y),
        'mortality': next(r['value'] for r in mortality if r['year'] == y),
    })

# downsample for chart legibility
sample_years = [y for y in years if y % 3 == 0 or y == years[-1]]
melanoma_trend = [r for r in melanoma_trend if r['year'] in sample_years]

# --- ABS behaviours ---
wb = load_workbook(DATA_DIR / 'SPBDC01.xlsx', read_only=True, data_only=True)

def read_age_rows(sheet_name, value_col_idx):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    start = None
    for idx, row in enumerate(rows):
        if row[0] == 'Age group':
            start = idx + 1
            break
    out = []
    if start is None:
        return out
    for row in rows[start:start+6]:
        label = row[0]
        value = row[value_col_idx]
        if label and isinstance(value, (int, float)):
            out.append({'label': str(label), 'value': float(value)})
    return out

# Persons column for sunscreen-used proportion is col H => index 7
sunscreen_age = read_age_rows('Table 1b', 7)
# Persons column for experienced sunburn is col H => index 7
sunburn_age = read_age_rows('Table 2b', 7)
# Persons column for attempted suntan is col H => index 7
suntan_age = read_age_rows('Table 3b', 7)

payload = {
    'melanomaTrend': melanoma_trend,
    'behaviourByAge': {
        'sunscreenMostDays': sunscreen_age,
        'sunburnLastWeek': sunburn_age,
        'attemptedSuntan': suntan_age,
    },
    'clothingByLevel': {
        'low': [
            'Low UV, but long outdoor time still deserves sunglasses and backup sunscreen.',
            'A breathable tee is fine for short trips; add a light overshirt if you will stay out longer.',
            'Keep a hat handy for midday sun.'
        ],
        'moderate': [
            'Breathable long-sleeve shirt or overshirt with tighter weave.',
            'Bucket hat or wide-brim hat.',
            'SPF 50+ on exposed skin and sunglasses.'
        ],
        'high': [
            'UPF or tightly woven long sleeves.',
            'Wide-brim hat that covers face, ears, and neck.',
            'Longer shorts or lightweight pants if you are outdoors for a while.',
            'Shade breaks during peak UV.'
        ],
        'very-high': [
            'UPF-rated long sleeves with neck coverage.',
            'Wide-brim hat, sunglasses, and SPF 50+ together.',
            'Lightweight long pants or other fuller coverage.',
            'Reduce time in direct midday sun.'
        ],
        'extreme': [
            'Minimise outdoor exposure where possible.',
            'Use maximum-coverage clothing plus a wide-brim hat.',
            'SPF 50+, sunglasses, and shade are essential.',
            'Avoid prolonged outdoor activity at peak times.'
        ]
    },
    'sources': [
        'Australian Cancer Incidence and Mortality (data.gov.au / AIHW)',
        'ABS Sun protection behaviours, Nov 2023 to Feb 2024 (SPBDC01.xlsx)',
        'Live UV: Open-Meteo current forecast API'
    ]
}

OUT.write_text('window.SUN_DATA = ' + json.dumps(payload, ensure_ascii=False, indent=2) + ';\n', encoding='utf-8')
print(f'Wrote {OUT}')
